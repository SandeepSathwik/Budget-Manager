from tkinter import *
import tkinter as tk
import random
#import tkinter as Tk
window = Tk()
import xlrd
from openpyxl import load_workbook
import webbrowser

loginDet = r"C:\Users\nithi\OneDrive\Desktop\LoginDetails.xlsx"
loginWB = xlrd.open_workbook(loginDet)
stu = loginWB.sheet_by_index(0)
loginWB1 = load_workbook(loginDet)
loginSheet = loginWB1['Sheet1']


def in_stuSign():
    stuSign = Tk()
    a = Label(stuSign, text="Full Name", font="Courier 20").grid(row=0, column=0)
    c = Label(stuSign, text="Password", font="Courier 20").grid(row=1, column=0)
    b = Label(stuSign, text="Age", font="Courier 20").grid(row=2, column=0)

    def addStudent():
        name = full_name.get()
        psswd = password.get()
        age = age_entry.get()
        col = loginSheet.max_row + 1
        loginSheet['A' + str(col)] = name
        loginSheet['B' + str(col)] = psswd
        loginSheet['C' + str(col)] = age
        loginWB1.save(loginDet)

    full_name = Entry(stuSign)
    full_name.grid(row=0, column=1)
    password = Entry(stuSign, show='*')
    password.grid(row=1, column=1)
    age_entry = Entry(stuSign)
    age_entry.grid(row=2, column=1)

    a = random.randint(999, 9999)
    n = Label(stuSign, text=a, bg='grey')
    n.grid(row=3, column=0)
    b = Entry(stuSign).grid(row=3, column=1)

    submit_button = Button(stuSign, text="Submit", command=addStudent)
    submit_button.grid(row=4, columnspan=2)


"""def sighup():
    def capcha():
        inputValue = textBox.get("1.0", "end-1c")
        print(inputValue)
    window1 = Tk()
    window1.geometry('500x400')
    username = Label(window1, text="User name", bg='grey')
    username.grid(row=1, column=0)
    password = Label(window1, text="Password", bg='grey')
    password.grid(row=3, column=0)
    usernameentry = Entry(window1)
    usernameentry.grid(row=2, column=0)
    passwordentry = Entry(window1)
    passwordentry.grid(row=4, column=0)

    def in_stuSign():
        a = Label(stuSign_frame, text="Full Name", font="Courier 20").grid(row=0, column=0)
        c = Label(stuSign_frame, text="Password", font="Courier 20").grid(row=1, column=0)
        b = Label(stuSign_frame, text="Year", font="Courier 20").grid(row=2, column=0)

        def addStudent():
            name = full_name.get()
            psswd = password.get()
            year = year_entry.get()
            col = stu1.max_row + 1
            stu1['A' + str(col)] = name
            stu1['B' + str(col)] = psswd
            stu1['C' + str(col)] = year
            stuwb1.save(stuloc)

        full_name = Entry(stuSign)
        full_name.grid(row=0, column=1)
        password = Entry(stuSign_frame, show='*')
        password.grid(row=1, column=1)
        year_entry = Entry(stuSign_frame)
        year_entry.grid(row=2, column=1)

        submit_button = Button(stuSign_frame, text="Submit", command=addStudent)
        submit_button.grid(row=3, columnspan=2)

        main = Button(stuSign_frame, text="Back", command=in_adminCource)
        main.grid(row=4, column=1)"""

"""def cheek():
        if(a==b.get()):
            done = Label(window1, text="Done")
            done.grid(row=12,column=0)
    btn3=Button(window1,text="Test1")
    btn3.grid(row=7,column=0)
    #########################
    textBox = Text(window1, height=2, width=10)
    textBox.grid(row=6,column=0)
    buttonCommit = Button(window1, height=1, width=10, text="test2",
                          command=lambda: capcha())
    # command=lambda: retrieve_input() >>> just means do this when i press the button
    buttonCommit.grid(row=8,column=1)
    ##############################
    #capc=Text(root, height=1, width=5,bg='pink')
    #capc.grid(row=8,column=0)
    #capctest=Button(window1,text="Test",
    #                command=lambda: capcha())
    #capctest.grid(row=9,column=0)
    ##############################
    btn = Button(window1, text="Sign up", bg='red', fg='white')
    btn.grid(row=10,column=0)
    window1.mainloop()"""
#btn = Button(window, text="Quit", command=quit)
#btn.grid(row=0, column=0)
window.geometry('950x550')
window.title("Budget Manager")
empty = Label(window, text=" ", bg='grey')
photo = PhotoImage(file='thumbnail_large.png')
label = Label(window, image=photo)
label.grid(row=2, column=4)

photo1 = PhotoImage(file='you.png')
label1 = Label(window, image=photo1)
label1.grid(row=14, column=0)

# empty.grid(row=0,column=0)
lbl = Label(window, text="Welcome to Budget Manager", font=("Arial Bold", 27), fg='white', bg="grey")
# lbl.pack(side=TOP)
lbl.grid(row=0, column=0)
window.configure(background="grey")


# --------------------------------------------------------------------------------

def stuPage():
    window3 = Toplevel()
    #window.quit()
    #btn=Button(window3,text="Quit",command=quit)
    #btn.grid(row=0,column=0)
    count = 1
    while (count < stu.nrows):
        if (usernameentry.get() == stu.cell_value(count, 0)):
            if (passwordentry.get() == stu.cell_value(count, 1)):
                def whatssapp():
                    window4 = Tk()
                    number = Label(window4, text="Whatss app phone number for any reviwe \n Phone Number:9876543210")
                    number.pack()
                    window4.mainloop()
                def OpenUrl1():
                    webbrowser.open_new('https://www.facebook.com/Budget-Manager-352149808880893/')

                def Openurl2():
                    webbrowser.open_new('https://twitter.com/login?lang=en')

                def Openurl3():
                    webbrowser.open_new('https://www.instagram.com/accounts/login/?hl=en')



                header = tk.Frame(window3, bg='lightgrey')
                content = tk.Frame(window3, bg='yellow')
                footer = tk.Frame(window3, bg='pink')

                window3.columnconfigure(0, weight=1)  # 100%
                window3.rowconfigure(0, weight=1)  # 10%
                window3.rowconfigure(1, weight=7)  # 70%
                window3.rowconfigure(2, weight=2)  # 20%

                header.grid(row=0, sticky='news')
                content.grid(row=1, sticky='news')
                footer.grid(row=2, sticky='news')
                ################

                ################
                window3.geometry('500x500')
                # photo4 = PhotoImage(file='profile.png')
                name = Label(header, text="NAME:", font=("Arial Bold", 27), fg='black', bg='lightgrey')
                name1 = Label(header, text=usernameentry.get(), font=("Arial Bold", 27,), fg='blue', bg='lightgrey')
                age = Label(header, text="Age:", font=("Arial Bold", 27), bg='lightgrey')
                age1 = Label(header, text="18", font=("Arial Bold", 27), bg='lightgrey')
                name.grid(row=0, column=1, sticky=N)
                name1.grid(row=0, column=2, sticky=N)
                age.grid(row=1, column=1, sticky=N)
                age1.grid(row=1, column=2, sticky=N)
                ##########################################################################################################
                ##########################################################################################################
                # content#
                def fuck():
                    test=Label(content,text="fuck",font=("Arial Bold", 10),bg="yellow")
                    test.grid(row=2,column=0,pady=10)
                text10=Label(content,text="   Click the calculator Button to start the program",bg='yellow', font=("Arial Bold", 15), fg='Black')
                text10.grid(row=0,column=0)
                btn10=Button(content,command=fuck)
                photo10 = PhotoImage(file='calc.png')
                btn10.config(image=photo10, width="50", height="50")
                btn10.grid(row=1, column=0)
                ##########################################################################################################















                ##########################################################################################################
                ##########################################################################################################

                photo4 = PhotoImage(file='profile.png')
                label4 = Label(header, image=photo4, bg='black')
                label4.grid(row=0, rowspan=2, column=0)

                logo1 = Button(footer, justify=LEFT, command=whatssapp)
                logo2 = Button(footer, justify=LEFT, command=OpenUrl1)
                logo3 = Button(footer, justify=LEFT, command=Openurl3)
                logo4 = Button(footer, justify=LEFT, command=Openurl2)

                photo11 = PhotoImage(file='whatsapp.png')
                photo12 = PhotoImage(file='facebook.png')
                photo13 = PhotoImage(file='instagram.png')
                photo14 = PhotoImage(file='twitter.png')

                logo1.config(image=photo11, width="50", height="50")
                logo2.config(image=photo12, width="50", height="50")
                logo3.config(image=photo13, width="50", height="50")
                logo4.config(image=photo14, width="50", height="50")
                txtt = Label(footer, text="Contact us at", bg="pink",font=("Helvetica",12))
                txtt.grid(row=0, column=0)
                logo1.grid(row=1, column=3, padx=10)
                logo2.grid(row=1, column=0, padx=10)
                logo3.grid(row=1, column=1, padx=10)
                logo4.grid(row=1, column=2, padx=10)
                window3.mainloop()
                ###################3
        count += 1

username = Label(window, text="User name", bg='grey')
username.grid(row=1, column=0)
password = Label(window, text="Password", bg='grey')
password.grid(row=1, column=1)
usernameentry = Entry(window)
usernameentry.grid(row=2, column=0)
passwordentry = Entry(window, show='*')
passwordentry.grid(row=2, column=1)
btn = Button(window, text="Sign in", bg='red', fg='white', command=stuPage)
btn.grid(row=2, column=2)
empty.grid(row=3)
text1 = Label(window, text="It would be Better if someone helps you to monitet your investments and ", fg='blue',
              bg='grey')
text2 = Label(window, text="organise your way of living and plan budget for you.", fg='blue', bg='grey')
text3 = Label(window, text="How about we do that for you?", fg='blue', bg='grey')
text4 = Label(window, text="To join create an account in budget manager there you can also", fg='blue', bg='grey')
text5 = Label(window, text="communicate with others and get suggetions. ", fg='blue', bg='grey')
text1.grid(row=5)
text2.grid(row=6)
text3.grid(row=7)
text4.grid(row=8)
text5.grid(row=9)
empty.grid(row=10)
text6 = Label(window, text="Create your account here", font=("Arial Bold", 17), fg='cyan', bg="grey")
text6.grid(row=11)
btn2 = Button(window, text="Sigh Up", fg='white', bg='red', command=in_stuSign)
btn2.grid(row=11, column=2)
empty.grid(row=12)
bottomframe = Label(window, bd=1, relief=SUNKEN, bg='darkgrey')
# bottomframe.grid(row=13)
# footer = tk.Frame(root, bg='green')
bottomframe.grid(row=13, column=0, sticky='news')
bottomframe.grid(row=13, column=1, sticky='news')
# username.grid_rowconfigure(0)
#cheek=Label(window,text="Test").grid(row=1,column=0)

username.grid_columnconfigure(0)
username.grid_rowconfigure(1)
username.grid_columnconfigure(0)

window.mainloop()